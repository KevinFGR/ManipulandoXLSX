# -*- coding: UTF-8 -*-
from openpyxl import Workbook
from openpyxl.styles import Font,PatternFill,Border,Side,Alignment
def main():
    # Função principal, enquanto não receber 'sair' continuará rodando.
    # 'mes' refere-se ao mês em que os dados da planilha se referem.
    # 'arq' refere-se ao arquivo dentro da pasta 'documentos', esta variável não precisa
    # da extensão '.txt'para entender e não deve ser adicionado o endereço 'documentos/' . 
    while True:
        mes = input('Digite o mês:')
        if mes.lower()=='sair': break
        arq = input('Digite o nome do arquivo:')
        if arq.lower() == 'sair': break
        arq = txt(arq)
        try:
            criaPlan(arq,mes)
            fim = input('\nPlanilha criada com sucesso.\n\nPressione "Enter" para criar outra planilha ou digite "sair"')
            if fim.lower() =='sair': break 
        except:
            print("ops!!! Houve um erro ao Processar dados\n")
            break

def txt(arq):
    #Adiciona a exenção caso o usuário não tenha digitado.
    if arq.endswith('.txt'): return arq
    else: return arq+'.txt'

def criaPlan(arq, mes):
    #Cria a planilha e suas primeiras linhas de titulos.
    nomeArq = arq.split('.')[0]
    wb = Workbook()
    plan =wb.active
    plan.title = mes
    plan['A1']=plan.title.capitalize()
    plan['A2']='Nome'
    plan['B2']='Seções'
    plan['C2']='Convenio'
    plan['D2']='Total'

    plan.merge_cells('A1:D1')
    plan['A1'].alignment = Alignment(horizontal='center', vertical='center')
    celulas = ['A1','A2','B2','C2','D2']
    estiliza(plan,celulas,1)

    pessoas = usaDoc(f'documentos/{arq}')
    addPessoas(pessoas, plan, 3)
    wb.save(f'planilhas/{nomeArq}.xlsx')

def estiliza(planilha,celulas,estilo):
    # Adiciona estilo às celulas recebidas em lista.
    # Caso estilo for 2, a celula assumirá estilo de título com background escuro e fontes claras.
    # Caso estilo for 1, a celula assumirá estilo de dados com background claro e fontes escuras.
    if estilo == 1:
        cor = 'ffffff'
        corFun = '1A5529'
        b = True
    elif estilo ==2:
        cor ='000000'
        corFun = 'D1F1DA'
        b = False
    else: return 0
    for i in celulas:
        planilha[i].font = Font(bold=b,color=cor)
        planilha[i].fill = PatternFill('solid', fgColor=corFun)
        larg = Side(border_style='thin',color='000000')
        planilha[i].border = Border(top=larg,bottom=larg,right=larg,left=larg)


def addPessoas(pessoas, plan, linhaInicial):
    #Adiciona as pessoas e seus dados na planilha.
    linha = linhaInicial
    for pessoa in pessoas:
        plan[f'A{linha}']=pessoa['nome']
        plan[f'B{linha}']=pessoa['quantidade']
        plan[f'C{linha}']=pessoa['convenio']
        plan[f'D{linha}']=pessoa['total']
        celulas = [f'A{linha}',f'B{linha}',f'C{linha}',f'D{linha}']
        estiliza(plan,celulas,2)
        linha +=1
    linhaFinal = [f'A{linha}',f'B{linha}',f'C{linha}',f'D{linha}']
    for i in linhaFinal:
        plan[i] = ' '
    estiliza(plan,linhaFinal,1)
    totalSomado = f'=SOMA(D{linhaInicial}:D{linha-1})'
    estiliza(plan,[f'D{linha+1}'],2)
    plan[f'D{linha+1}']=totalSomado

def temRepetidos(lista,nome):
    # Verifica se o nome já está na lista. Caso esteja, irá retornar uma lista com verdadeiro no
    # primeiro indice e no segundo indice retornará o índice na lista onde o nome que já existe está.
    for i in lista:
        if i['nome'] == nome:
            return [True,[i]]
    return [False]


def usaDoc(documento):
    #Abre o documento e acha as linhas que estão com os dados necessarios.
    #Retorna uma lista de dicionarios com os dados.
    with open(documento, 'r', encoding="UTF-8") as arq:
        tudo = arq.read()
        blocos = tudo.split('Psicologia')[1:]
        retorno = []
        
        for bloco in blocos:
            linhas = bloco.strip().split("\n")
            convenio = linhas[1].strip()[1:].split('.')[0].strip()
            for linha in linhas: 
                if 'Nome' in linha:
                    nome = pegaNome(linha)
                    if 'Quantidade' in linha:
                        qtd = pegaQtd(linha)
                elif 'Quantidade' in linha:
                    qtd = pegaQtd(linha)  
            if retorno != []: 
                ultimoAdd = retorno[-1]
                if ultimoAdd['nome']==nome:
                    ultimoAdd['quantidade']+=qtd
                    total = precoConvenio(ultimoAdd['convenio'],ultimoAdd['quantidade'])
                else:
                    total = precoConvenio(convenio,qtd)
                    pessoa = {'nome': nome, 'convenio':convenio, 'quantidade':qtd, 'total':total}
                    retorno.append(pessoa)
            else:                   
                total = precoConvenio(convenio,qtd)
                pessoa = {'nome': nome, 'convenio':convenio, 'quantidade':qtd, 'total':total}
                retorno.append(pessoa)
        return retorno

def precoConvenio(convenio, qtd):
    #Retorna o valor total referente ao convenio da pessoa e a qtd de seções.
    qtd = float(qtd)
    convenio = convenio.lower()
    if 'sul' in convenio: return qtd*27.56
    elif 'amil' in convenio: return qtd*24.39
    elif 'notre' in convenio: return qtd*15.00
    elif 'porto' in convenio: return qtd*16.34
    elif 'sompo' in convenio: return  qtd*9.67
    elif 'particular' in convenio: return qtd*35.00
    elif 'unimed' in convenio: return qtd*12.00
    else: return 0

def pegaQtd(frase):
    #Garimpa a penas o numero referente à quantidade na linha.
    qtd= frase[frase.find('Quantidade'):].split(':')[1].split('.')[0].split()[0]
    return int(qtd)

def pegaNome(frase):
    #Retorna apenas o nome,da pessoa, que está dentro da linha.
    nome = frase[frase.find('Nome'):].split(':')[1].split('.')[0].strip()
    return nome
    
if __name__ == '__main__':
    # Caso este arquivo seja executado diretamente, a função main será iniciada
    main()
